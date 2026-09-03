# -*- coding: utf-8 -*-
"""
项目导出 Excel 模块
"""

import os
from datetime import datetime

from PyQt5.QtWidgets import QFileDialog, QMessageBox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from models import short_datetime


def export_projects_to_xlsx(data, parent=None):
    """导出项目数据到 Excel"""
    if not data.projects:
        QMessageBox.information(parent, "导出", "暂无项目可导出")
        return

    default_name = f"项目看板_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    path, _ = QFileDialog.getSaveFileName(
        parent, "导出项目看板", default_name, "Excel 文件 (*.xlsx)"
    )
    if not path:
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "项目看板"

        # 表头
        headers = ["序号", "项目名称", "状态", "负责人", "开始时间", "截止日期", "待协调事项", "备注", "创建时间", "最后更新时间"]
        ws.append(headers)

        # 表头样式
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 内容
        sorted_projects = data.sorted_projects()
        status_colors = {
            "未开始": "D1D5DB",
            "进行中": "BFDBFE",
            "待协调": "FECACA",
            "已完成": "BBF7D0",
            "暂停": "E5E7EB",
        }

        for idx, p in enumerate(sorted_projects, start=1):
            row = [
                idx,
                p.get("name", ""),
                p.get("status", ""),
                p.get("owner", "") or "-",
                p.get("startDate", "") or "-",
                p.get("deadline", "") or "-",
                p.get("pending", "") or "-",
                p.get("remark", "") or "-",
                short_datetime(p.get("createdAt", "")),
                short_datetime(p.get("updatedAt", "")),
            ]
            ws.append(row)

            # 状态列上色
            status_cell = ws.cell(row=idx + 1, column=3)
            status_cell.fill = PatternFill(
                start_color=status_colors.get(p.get("status", ""), "FFFFFF"),
                end_color=status_colors.get(p.get("status", ""), "FFFFFF"),
                fill_type="solid"
            )
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 设置列宽
        col_widths = [8, 30, 12, 15, 15, 15, 35, 25, 20, 20]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[__import__("openpyxl").utils.get_column_letter(i)].width = width

        # 所有单元格加边框
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 冻结首行
        ws.freeze_panes = "A2"

        wb.save(path)
        QMessageBox.information(parent, "导出成功", f"已成功导出到:\n{path}")
    except Exception as e:
        QMessageBox.critical(parent, "导出失败", f"导出 Excel 失败:\n{str(e)}")
